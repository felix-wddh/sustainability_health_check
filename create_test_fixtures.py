"""
create_test_fixtures.py - Generate synthetic XLSX test files with known values.

Run this script to create test fixtures in the fixtures/ directory.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


def create_dryer_smg():
    """Create a Dryer SMG workbook with anchor-based energy values."""
    wb = Workbook()
    
    # Sheet 1: Cover/Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Mabe Product Sustainability Report"
    ws1["A2"] = "Product: Dryer SMG (SMG6527XXXX)"
    ws1["A3"] = "Date: 2025-01-15"
    
    # Sheet 2: Model Data (anchor-based format)
    ws2 = wb.create_sheet("Dryer SMG (SMG6527)")
    ws2["A1"] = "Product Specifications"
    ws2["A3"] = "Model Number"
    ws2["B3"] = "SMG6527XXXX"
    ws2["A5"] = "Annual Energy Consumption"
    ws2["B5"] = "409.6 kWh"
    ws2["A7"] = "Transport CO2"
    ws2["B7"] = 4.5
    ws2["A9"] = "Materials CO2"
    ws2["B9"] = 85.2
    ws2["A11"] = "Production CO2"
    ws2["B11"] = 18.3
    
    # Sheet 3: Another model for comparison
    ws3 = wb.create_sheet("Dryer GTD (GTD42XXX)")
    ws3["A1"] = "Product Specifications"
    ws3["A3"] = "Model Number"
    ws3["B3"] = "GTD42XXXX"
    ws3["A5"] = "Annual Energy Consumption"
    ws3["B5"] = "245.0 kWh/year"
    ws3["A7"] = "Transport CO2"
    ws3["B7"] = 3.8
    ws3["A9"] = "Materials CO2"
    ws3["B9"] = 72.1
    ws3["A11"] = "Production CO2"
    ws3["B11"] = 15.6
    
    filepath = os.path.join(FIXTURES_DIR, "dryer_workbook.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath


def create_refrigerator():
    """Create a Refrigerator workbook with different energy values."""
    wb = Workbook()
    
    # Sheet 1: Cover
    ws1 = wb.active
    ws1.title = "Cover"
    ws1["A1"] = "Refrigeration Product Assessment"
    
    # Sheet 2: Model Data (anchor-based)
    ws2 = wb.create_sheet("Cooling Unit (GSS25XXX)")
    ws2["A1"] = "Refrigerator Specifications"
    ws2["A3"] = "SKU"
    ws2["B3"] = "GSS25XXXX"
    ws2["A5"] = "Energy Consumption"
    ws2["B5"] = 322
    ws2["C5"] = "kWh/year"
    ws2["A7"] = "Logistics"
    ws2["B7"] = 6.2
    ws2["A9"] = "BOM"
    ws2["B9"] = 120.5
    ws2["A11"] = "Manufacturing"
    ws2["B11"] = 22.8
    
    filepath = os.path.join(FIXTURES_DIR, "refrigerator_workbook.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath


def create_table_format():
    """Create a workbook with table-format data (header row + data rows)."""
    wb = Workbook()
    
    ws = wb.active
    ws.title = "Products Table"
    
    # Header row
    headers = ["Product", "Category", "Transport kgCO2e", "Materials kgCO2e", 
               "Production kgCO2e", "kWh per year"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data rows
    data = [
        ["WTW5000DW", "Washing", 2.5, 65.0, 12.0, 175.5],
        ["WTW5105HW", "Washing", 2.8, 68.5, 13.2, 168.0],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    filepath = os.path.join(FIXTURES_DIR, "washer_table_format.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath


def create_spanish_format():
    """Create a workbook with Spanish labels."""
    wb = Workbook()
    
    ws = wb.active
    ws.title = "Secadora (SMG1234)"
    
    ws["A1"] = "Especificaciones del Producto"
    ws["A3"] = "Modelo"
    ws["B3"] = "SMG1234"
    ws["A5"] = "Consumo de energía"
    ws["B5"] = "385,5"  # European comma decimal
    ws["C5"] = "kWh/año"
    ws["A7"] = "Transporte"
    ws["B7"] = "5,2"
    ws["A9"] = "Materiales"
    ws["B9"] = "92,3"
    ws["A11"] = "Producción"
    ws["B11"] = "19,8"
    
    filepath = os.path.join(FIXTURES_DIR, "secadora_spanish.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")
    return filepath


def main():
    """Create all test fixtures."""
    os.makedirs(FIXTURES_DIR, exist_ok=True)
    
    create_dryer_smg()
    create_refrigerator()
    create_table_format()
    create_spanish_format()
    
    print(f"\nAll fixtures created in: {FIXTURES_DIR}")


if __name__ == "__main__":
    main()
